import pandas as pd
import os
from openpyxl import Workbook

def judgment_grade(data):
    
    match data['grade']:
        case '一年級':
            if data['score'] >= 50:
                return '金質獎'
            elif data['score'] >= 40:
                return '銀質獎'
            elif data['score']>= 30:
                return '銅質獎'
            else:
                return None
        case '二年級':
            if data['score'] >= 80:
                return '金質獎'
            elif data['score'] >= 60:
                return '銀質獎'
            elif data['score']>= 40:
                return '銅質獎'
            else:
                return None
        case '三年級':
            if data['score'] >= 95:
                return '金質獎'
            elif data['score'] >= 90:
                return '銀質獎'
            elif data['score']>= 85:
                return '銅質獎'
            else:
                return None
        case '四年級':
            if data['score'] >= 100:
                return '金質獎'
            elif data['score'] >= 95:
                return '銀質獎'
            elif data['score']>= 90:
                return '銅質獎'
            else:
                return None
        case '五年級':
            if data['score'] >= 50:
                return '金質獎'
            elif data['score'] >= 45:
                return '銀質獎'
            elif data['score']>= 40:
                return '銅質獎'
            else:
                return None
        case '六年級':
            if data['score'] >= 55:
                return '金質獎'
            elif data['score'] >= 50:
                return '銀質獎'
            elif data['score']>= 45:
                return '銅質獎'
            else:
                return None
        case '推廣組':
            if data['score'] >= 70:
                return '金質獎'
            elif data['score'] >= 50:
                return '銀質獎'
            elif data['score']>= 30:
                return '銅質獎'
            else:
                return None
        case '特教組':
            if data['score'] >= 70:
                return '金質獎'
            elif data['score'] >= 50:
                return '銀質獎'
            elif data['score']>= 30:
                return '銅質獎'
            else:
                return None
    
def student():
    df = pd.read_json('..\\jumpBackend\\contestants.json')
    
    #print(len(df))
    for i in range( len(df) ):
        k = judgment_grade(df.loc[i,:])
        df.loc[i , '等第'] = k
    
    
    df.rename (columns={"contest":"項目","group":"組別", "grade":"年級", "school":"學校", "name":"姓名", "teacher":"指導老師" , "id":"編號","score":"成績","等第":"等第" },inplace=True)    
    
    df.to_excel('..\\名單\\112跳繩各校學生成績.xlsx',index=False)
    
    df.drop(columns=['組別','成績','編號'])
    
    new_order = ['年級', '學校', '項目', '等第', '姓名', '指導老師']
    df = df.reindex(columns=new_order)
    df.columns = ['參賽組別', '單位', '項目', '成績', '優秀選手', '指導老師']
    
    df.dropna(subset=['成績'], inplace=True)
    df['項目']=['單人繩' for i in range(len(df))]    
    df['參賽組別'] = df['參賽組別'].apply(lambda x: x + '組')     
    
    df.to_excel('..\\名單\\112跳繩各校學生獎狀.xlsx',index=False)
    
def teacher():
    
    df = pd.read_excel('..\\名單\\112跳繩各校學生成績.xlsx')
    
    # 找出所有等第非None的row
    df = df.dropna(subset=['等第'])

    # 按指導老師分組
    groups = df.groupby(['指導老師'])

    new_rows = []

    # 遍歷每個指導老師的分組
    for name, group in groups:
        # 按年級和組別分出不同年級和男女組
        grade_groups = group.groupby(['年級', '組別'])
        
        # 選擇每個年級和組別中的最高成績
        for gname, ggroup in grade_groups:
            max_row = ggroup.loc[ggroup['成績'].idxmax()]
            new_rows.append(max_row)

    # 建立新的dataframe
    new_df = pd.DataFrame(new_rows, columns=df.columns)
    
    # 製作輸出格式
    new_df.drop(columns=['組別','成績','編號'])
    new_order = ['年級', '學校', '項目', '等第', '姓名', '指導老師']
    new_df = new_df.reindex(columns=new_order)
    new_df.columns = ['參賽組別', '單位', '項目', '成績', '優秀選手', '指導老師']
    new_df['項目']=['單人繩' for i in range(len(new_df))]    
    new_df['參賽組別'] = new_df['參賽組別'].apply(lambda x: x + '組') 
    
    new_df.to_excel('..\\名單\\112跳繩各校老師獎狀.xlsx',index=False)

def make_sheet():
    # 讀取資料並分組
    df = pd.read_excel('..\\名單\\112跳繩各校學生成績.xlsx')
   
    groups = df.groupby('學校')

    # 建立新的 Excel 檔案
    wb = Workbook()

    # 將每個分組寫入不同工作表中
    for school, group in groups:
        sheet_name = f'{school[0]+school[1]}'
        print(school[0]+school[1])
        ws = wb.create_sheet(sheet_name)
        ws.append(group.columns.tolist())
        for row in group.iterrows():
            ws.append(list(row[1]))

    if 'Sheet' in wb.sheetnames:
        sheet = wb['Sheet']
        wb.remove(sheet)

    wb.save('..\\名單\\112跳繩各校學生成績-分類版.xlsx')

if __name__ == '__main__':
    student()
    teacher()
    make_sheet()
    
    
